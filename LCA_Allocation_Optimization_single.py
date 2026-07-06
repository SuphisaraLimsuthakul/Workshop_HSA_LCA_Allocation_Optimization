from __future__ import annotations
from collections import defaultdict
import math
import re
import time
import datetime
import pandas as pd
from ortools.linear_solver import pywraplp
from openpyxl.styles import Alignment, Font, PatternFill


# ----------------------------
# Utils
# ----------------------------
def _ordered_unique(values):
    seen = set()
    out = []
    for v in values:
        if pd.isna(v):
            continue
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _safe_sheet_name(name: str) -> str:
    text = str(name)
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        text = text.replace(ch, "_")
    text = text.strip() or "Sheet"
    return text[:31]


def _ensure_month_key(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "MonthKey" in out.columns:
        out["MonthKey"] = out["MonthKey"].astype(str).str.strip()
        return out
    if {"Year", "Quarter", "Month"}.issubset(out.columns):
        out["MonthKey"] = out["Year"].astype(str).str.strip() + "_" + out["Quarter"].astype(str).str.strip() + "_" + out["Month"].astype(str).str.strip()
        return out
    if "Month" in out.columns:
        out["MonthKey"] = out["Month"].astype(str).str.strip()
        return out
    raise ValueError("Input must contain either MonthKey, or Year+Quarter+Month, or Month column.")


def _prepare_idle_month_key(df_idle: pd.DataFrame, month_keys: list[str]) -> pd.DataFrame:
    out = df_idle.copy()
    if ("MonthKey" in out.columns) or ({"Year", "Quarter", "Month"}.issubset(out.columns)) or ("Month" in out.columns):
        return _ensure_month_key(out)

    if "Product" not in out.columns:
        raise ValueError("df_idel_final must contain 'Product' column.")

    value_col = "Value" if "Value" in out.columns else None
    if value_col is None:
        out["Value"] = 0.0
        value_col = "Value"

    idle_by_product = out.groupby("Product", dropna=True)[value_col].sum().reset_index()
    expanded_rows = []
    for _, row in idle_by_product.iterrows():
        for mk in month_keys:
            expanded_rows.append({"Product": row["Product"], "MonthKey": mk, "Value": float(row[value_col])})
    return pd.DataFrame(expanded_rows)


def _build_month_labels(df_req: pd.DataFrame, month_keys: list[str]) -> dict[str, str]:
    label_by_key: dict[str, str] = {}

    if {"MonthKey", "Year", "Quarter", "Month"}.issubset(df_req.columns):
        base = (
            df_req[["MonthKey", "Year", "Quarter", "Month"]]
            .dropna(subset=["MonthKey"])
            .drop_duplicates(subset=["MonthKey"], keep="first")
        )
        for _, row in base.iterrows():
            mk = str(row["MonthKey"]).strip()
            year_val = row["Year"]
            if pd.isna(year_val):
                year_text = ""
            elif isinstance(year_val, (int, float)) and float(year_val).is_integer():
                year_text = str(int(year_val))
            else:
                year_text = str(year_val).strip()
            quarter = str(row["Quarter"]).strip()
            month = str(row["Month"]).strip()
            if quarter and month and year_text:
                label_by_key[mk] = f"{quarter}-{year_text}-{month}"
            elif quarter and month:
                label_by_key[mk] = f"{quarter}-{month}"
            elif month:
                label_by_key[mk] = month

    if not label_by_key and {"MonthKey", "Quarter", "Month"}.issubset(df_req.columns):
        base = (
            df_req[["MonthKey", "Quarter", "Month"]]
            .dropna(subset=["MonthKey"])
            .drop_duplicates(subset=["MonthKey"], keep="first")
        )
        for _, row in base.iterrows():
            mk = str(row["MonthKey"]).strip()
            quarter = str(row["Quarter"]).strip()
            month = str(row["Month"]).strip()
            if quarter and month:
                label_by_key[mk] = f"{quarter}-{month}"
            elif month:
                label_by_key[mk] = month

    if not label_by_key and {"MonthKey", "Month"}.issubset(df_req.columns):
        base = (
            df_req[["MonthKey", "Month"]]
            .dropna(subset=["MonthKey"])
            .drop_duplicates(subset=["MonthKey"], keep="first")
        )
        for _, row in base.iterrows():
            mk = str(row["MonthKey"]).strip()
            label_by_key[mk] = str(row["Month"]).strip()

    for mk in month_keys:
        label_by_key.setdefault(mk, mk)

    return label_by_key


def _strip_col(df: pd.DataFrame, col: str) -> pd.DataFrame:
    out = df.copy()
    if col in out.columns:
        out[col] = out[col].astype(str).str.strip()
    return out


def _ceil_with_tolerance(value: float, tolerance: float = 1e-9) -> float:
    numeric_value = float(value)
    if numeric_value <= tolerance:
        return 0.0
    return float(math.ceil(numeric_value - tolerance))


def _get_display_inventory_value(
    product: str,
    month_idx: int,
    model_value: float,
    data: dict,
) -> float:
    # Preserve original decimal opening inventory in month 1 displays while
    # keeping integer model variables for optimization.
    if month_idx != 1:
        return float(model_value)
    opening_raw = float(data.get("A0Raw", {}).get(product, model_value))
    opening_model = float(data.get("A0", {}).get(product, model_value))
    return float(model_value) + (opening_raw - opening_model)



def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


_FLOW_COLOR_PALETTE = [
    "FDE2E4",
    "E2ECE9",
    "F9EAC2",
    "D9ED92",
    "B5EAD7",
    "C7CEEA",
    "E8DFF5",
    "FFD6A5",
    "CAE9FF",
    "D8F3DC",
]


_PREFERRED_PRODUCT_ORDER_ALIASES: list[set[str]] = [
    {"pharaohoasishepburnoasis", "pharaohoasishepburnoasis"},
    {"bacall4hd"},
    {"v15v15cmr24hd"},
    {"v154d"},
    {"v15cmrpi4d"},
    {"evansbp"},
    {"longspeakbp", "longspeaklongspeakbp"},
    {"marlinbpandsummit", "marlinmarlinbpsummit"},
    {"dorado10d5d", "dorado10ddorado5d"},
    {"bluewhale"},
    {"harrier"},
    {"osprey"},
    {"rosewood71d"},
    {"rosewood72d"},
    {"cimarronbp"},
]


def _preferred_product_rank(product_name: str) -> int:
    norm = _normalize_key(product_name)
    if not norm:
        return len(_PREFERRED_PRODUCT_ORDER_ALIASES)

    for idx, aliases in enumerate(_PREFERRED_PRODUCT_ORDER_ALIASES):
        if norm in aliases:
            return idx

    return len(_PREFERRED_PRODUCT_ORDER_ALIASES)


def _sort_products_for_output(products: list[str]) -> list[str]:
    original_index = {p: i for i, p in enumerate(products)}
    return sorted(
        products,
        key=lambda p: (_preferred_product_rank(p), original_index[p]),
    )


def _split_group_members(text: str) -> list[str]:
    raw = str(text).strip()
    if not raw:
        return []

    # Split only on top-level "/" so product names like "Dorado (10D/5D)" stay intact.
    parts = []
    current = []
    depth = 0
    for ch in raw:
        if ch == "(":
            depth += 1
            current.append(ch)
            continue
        if ch == ")":
            depth = max(depth - 1, 0)
            current.append(ch)
            continue
        if ch == "/" and depth == 0:
            piece = "".join(current).strip()
            if piece:
                parts.append(piece)
            current = []
            continue
        current.append(ch)

    tail = "".join(current).strip()
    if tail:
        parts.append(tail)

    expanded = []
    for part in parts:
        and_split = re.split(r"\s+and\s+", part, flags=re.IGNORECASE)
        for chunk in and_split:
            chunk = chunk.strip()
            if chunk:
                expanded.append(chunk)

    return _ordered_unique(expanded)


def _build_group_alias_map(products: list[str], group_names: list[str]) -> dict[str, set[str]]:
    alias_to_products: dict[str, set[str]] = defaultdict(set)

    for p in products:
        aliases = [p] + _split_group_members(p)
        for alias in aliases:
            alias_norm = _normalize_key(alias)
            if alias_norm:
                alias_to_products[alias_norm].add(p)

    # Expand aliases from explicit group sheet to all product names in that group.
    for group_name in group_names:
        members = _split_group_members(group_name)
        if not members:
            continue

        group_products = set()
        for member in members:
            member_norm = _normalize_key(member)
            if member_norm and member_norm in alias_to_products:
                group_products.update(alias_to_products[member_norm])

        if not group_products:
            continue

        for member in members:
            member_norm = _normalize_key(member)
            if member_norm:
                alias_to_products[member_norm].update(group_products)

    return alias_to_products


def _resolve_products_from_alias(name: str, products: list[str], alias_map: dict[str, set[str]]) -> list[str]:
    norm_name = _normalize_key(name)
    if not norm_name:
        return []

    exact = alias_map.get(norm_name, set())
    if exact:
        return [p for p in products if p in exact]

    # Fallback for minor formatting differences.
    matched = [p for p in products if norm_name in _normalize_key(p) or _normalize_key(p) in norm_name]
    return _ordered_unique(matched)


def _cost_family_key(name: str) -> str | None:
    norm_name = _normalize_key(name)
    if "bluewhale" in norm_name:
        return "bluewhale"
    if "dorado" in norm_name:
        return "dorado"
    return None


def _expand_products_by_cost_family(mapped_products: list[str], products: list[str]) -> list[str]:
    if not mapped_products:
        return []

    expanded = set(mapped_products)
    families = {_cost_family_key(p) for p in mapped_products}
    families.discard(None)

    if not families:
        return mapped_products

    for p in products:
        if _cost_family_key(p) in families:
            expanded.add(p)

    return [p for p in products if p in expanded]


def _build_idle_active_by_month(
    products: list[str],
    months: list[int],
    requirements: dict[tuple[str, int], float],
    base_idle_products: list[str],
    zero_run_threshold: int = 3,
) -> dict[tuple[str, int], bool]:
    idle_active: dict[tuple[str, int], bool] = {}
    base_idle_set = set(base_idle_products)

    for p in products:
        zero_run_start_idx = None
        zero_run_length = 0

        def _flush_zero_run_if_needed(end_idx_exclusive: int) -> None:
            nonlocal zero_run_start_idx, zero_run_length
            if zero_run_start_idx is None:
                return
            if zero_run_length >= zero_run_threshold:
                for run_idx in range(zero_run_start_idx, end_idx_exclusive):
                    t_run = months[run_idx]
                    idle_active[(p, t_run)] = True
            zero_run_start_idx = None
            zero_run_length = 0

        for t in months:
            idle_active[(p, t)] = (p in base_idle_set)

        for idx, t in enumerate(months):
            req_value = float(requirements.get((p, t), 0.0))
            if abs(req_value) <= 1e-9:
                if zero_run_start_idx is None:
                    zero_run_start_idx = idx
                zero_run_length += 1
            else:
                _flush_zero_run_if_needed(idx)

        _flush_zero_run_if_needed(len(months))

    return idle_active


def _solver_status_name(status: int) -> str:
    status_map = {
        pywraplp.Solver.OPTIMAL: "OPTIMAL",
        pywraplp.Solver.FEASIBLE: "FEASIBLE",
        pywraplp.Solver.INFEASIBLE: "INFEASIBLE",
        pywraplp.Solver.UNBOUNDED: "UNBOUNDED",
        pywraplp.Solver.ABNORMAL: "ABNORMAL",
        pywraplp.Solver.NOT_SOLVED: "NOT_SOLVED",
    }
    return status_map.get(status, f"UNKNOWN({status})")


def _diagnose_infeasibility(data: dict, tolerance: float = 1e-9) -> list[str]:
    """Return likely infeasibility causes based on model hard constraints."""
    issues: list[str] = []
    products = data.get("Products", [])
    months = data.get("Months", [])
    requirements = data.get("R", {})
    opening_inventory = data.get("A0", {})

    if len(months) >= 1:
        m1 = months[0]
        for p in products:
            r1 = float(requirements.get((p, m1), 0.0))
            a0 = float(opening_inventory.get(p, 0.0))
            if r1 - a0 > tolerance:
                issues.append(
                    f"Month {m1}: requirement > opening inventory for '{p}' (R={r1}, A0={a0}). "
                    "Model forces Short=0 in first month."
                )

    if len(months) >= 2:
        m2 = months[1]
        for p in products:
            r2 = float(requirements.get((p, m2), 0.0))
            a0 = float(opening_inventory.get(p, 0.0))
            if r2 - a0 > tolerance:
                issues.append(
                    f"Month {m2}: requirement > available pre-buy inventory for '{p}' (R={r2}, A0={a0}). "
                    "Model forces Short=0 in second month and blocks early conversion."
                )

    if not issues:
        issues.append(
            "No quick pre-check violation found. Likely conflict is from combined inventory/"
            "conversion timing constraints. Export LP and inspect infeasible rows."
        )

    return issues


# ----------------------------
# 1) Load & Prepare Input (Demo_input.xlsx)
# ----------------------------
def build_data_from_demo_input(
    path: str,
):
    df_req = pd.read_excel(path, sheet_name="df_re_final")
    df_idle = pd.read_excel(path, sheet_name="df_idel_final")
    df_inv = pd.read_excel(path, sheet_name="df_in_final")
    df_conv_cost = pd.read_excel(path, sheet_name="df_con_cost")
    df_cost = pd.read_excel(path, sheet_name="df_cost")
    try:
        df_group = pd.read_excel(path, sheet_name="df_group")
    except Exception:
        df_group = pd.DataFrame(columns=["Group Name"])

    # Normalize product keys for reliable joins/lookups
    df_req = _strip_col(df_req, "Product")
    df_idle = _strip_col(df_idle, "Product")
    df_inv = _strip_col(df_inv, "Product")
    df_conv_cost = _strip_col(df_conv_cost, "Product Name (From)")
    df_conv_cost = _strip_col(df_conv_cost, "Product Name (To)")
    df_cost = _strip_col(df_cost, "Product")

    # Guard against NaN/non-numeric costs that can make CBC return ABNORMAL (status=4).
    df_conv_cost["Cost"] = pd.to_numeric(df_conv_cost["Cost"], errors="coerce")
    df_conv_cost = df_conv_cost[df_conv_cost["Cost"].notna()].copy()
    df_cost["Cost"] = pd.to_numeric(df_cost["Cost"], errors="coerce")
    df_cost = df_cost[df_cost["Cost"].notna()].copy()

    # Month mapping
    df_req = _ensure_month_key(df_req)
    month_keys = _ordered_unique(df_req["MonthKey"].tolist())
    month_label_by_key = _build_month_labels(df_req, month_keys)
    df_idle = _prepare_idle_month_key(df_idle, month_keys)

    Months = list(range(1, len(month_keys) + 1))
    month_map = dict(zip(month_keys, Months))
    month_label_by_index = {
        month_map[mk]: month_label_by_key.get(mk, mk)
        for mk in month_keys
    }

    products = _ordered_unique(
        list(df_req["Product"].tolist())
        + list(df_idle["Product"].tolist())
        + list(df_inv["Product"].tolist())
    )
    products = _sort_products_for_output(products)
    idle_products = _ordered_unique(df_idle["Product"].dropna().tolist())
    group_names = []
    if "Group Name" in df_group.columns:
        group_names = _ordered_unique(df_group["Group Name"].dropna().astype(str).str.strip().tolist())

    # Requirement (model)
    req_grp = df_req.groupby(["Product", "MonthKey"], dropna=True)["Value_Round"].sum().reset_index()
    R = {(row["Product"], month_map[row["MonthKey"]]): float(row["Value_Round"]) for _, row in req_grp.iterrows()}
    for p in products:
        for t in Months:
            R.setdefault((p, t), 0.0)

    # Requirement display values for output sheets.
    # Prefer business actual value from input "Value".
    req_display_col = "Value" if "Value" in df_req.columns else ("Round2" if "Round2" in df_req.columns else "Value_Round")
    req_display_grp = (
        df_req.groupby(["Product", "MonthKey"], dropna=True)[req_display_col]
        .sum()
        .reset_index()
    )
    RRound2 = {
        (row["Product"], month_map[row["MonthKey"]]): float(row[req_display_col])
        for _, row in req_display_grp.iterrows()
    }
    for p in products:
        for t in Months:
            RRound2.setdefault((p, t), 0.0)

    # Initial inventory
    inv_grp = df_inv.groupby(["Product"], dropna=True)["Value"].sum().reset_index()
    A0 = {row["Product"]: float(row["Value"]) for _, row in inv_grp.iterrows()}
    for p in products:
        A0.setdefault(p, 0.0)

    # Idle capacity per product per month
    idle_grp = df_idle.groupby(["Product", "MonthKey"], dropna=True)["Value"].sum().reset_index()
    IdleCapProd = {
        (row["Product"], month_map[row["MonthKey"]]): float(row["Value"]) for _, row in idle_grp.iterrows()
    }
    for p in products:
        for t in Months:
            IdleCapProd.setdefault((p, t), 0.0)

    # Treat month-1 idle stock as available opening inventory.
    # This allows idle-only sources (e.g. EvansBP) to participate in conversion decisions.
    if Months:
        first_month = Months[0]
        for p in products:
            A0[p] = float(A0.get(p, 0.0)) + float(IdleCapProd.get((p, first_month), 0.0))

    A0Raw = {p: float(A0.get(p, 0.0)) for p in products}
    A0 = {p: _ceil_with_tolerance(A0Raw[p]) for p in products}

    # Business rule: if first-month requirement exceeds available opening inventory,
    # cap requirement down to opening inventory so month-1 allocation follows inventory.
    if Months:
        first_month = Months[0]
        for p in products:
            req_first = float(R.get((p, first_month), 0.0))
            inv_first = float(A0.get(p, 0.0))
            if req_first > inv_first:
                R[(p, first_month)] = inv_first

    IdleActiveByMonth = _build_idle_active_by_month(
        products=products,
        months=Months,
        requirements=R,
        base_idle_products=idle_products,
        zero_run_threshold=3,
    )

    # Conversion cost q->p
    conv_grp = df_conv_cost.groupby(["Product Name (From)", "Product Name (To)"], dropna=True)["Cost"].mean().reset_index()
    alias_map = _build_group_alias_map(products, group_names)
    pair_cost_values: dict[tuple[str, str], list[float]] = defaultdict(list)

    for _, row in conv_grp.iterrows():
        src_name = str(row["Product Name (From)"]).strip()
        dst_name = str(row["Product Name (To)"]).strip()
        cost_value = float(row["Cost"])
        if not pd.notna(cost_value):
            continue

        src_products = _resolve_products_from_alias(src_name, products, alias_map)
        dst_products = _resolve_products_from_alias(dst_name, products, alias_map)
        src_products = _expand_products_by_cost_family(src_products, products)
        dst_products = _expand_products_by_cost_family(dst_products, products)

        for q in src_products:
            for p in dst_products:
                if q != p:
                    pair_cost_values[(q, p)].append(cost_value)

    ConvCost = {pair: float(sum(vals) / len(vals)) for pair, vals in pair_cost_values.items()}
    product_set = set(products)
    idle_product_set = set(idle_products)
    AllowedConvPairs = {
        (q, p)
        for (q, p) in ConvCost.keys()
        if q in product_set and p in product_set and q != p
        and not (q in idle_product_set and p in idle_product_set)
        # and (q, p) not in _FORBIDDEN_CONVERSION_PAIRS
    }

    idle_eligible_products = [
        p for p in products
        if any(IdleActiveByMonth[(p, t)] for t in Months)
    ]

    missing_idle_conv_cost_products = [
        q for q in idle_products
        if not any((q, p) in AllowedConvPairs for p in products if p != q)
    ]
    missing_idle_conv_cost_products.extend(
        q for q in idle_eligible_products
        if q not in missing_idle_conv_cost_products
        and not any((q, p) in AllowedConvPairs for p in products if p != q)
    )

    # New Buy cost (optional): allow member-level names in df_cost to map to grouped products.
    buy_cost_values: dict[str, list[float]] = defaultdict(list)
    for _, row in df_cost.dropna(subset=["Product"]).iterrows():
        src_name = str(row["Product"]).strip()
        try:
            cost_value = float(row["Cost"])
        except Exception:
            continue
        if not pd.notna(cost_value):
            continue

        mapped_products = _resolve_products_from_alias(src_name, products, alias_map)
        mapped_products = _expand_products_by_cost_family(mapped_products, products)
        for p in mapped_products:
            buy_cost_values[p].append(cost_value)

    NewBuyCost = {
        p: float(sum(vals) / len(vals))
        for p, vals in buy_cost_values.items()
        if vals
    }
    for p in products:
        NewBuyCost.setdefault(p, 0.0)

    return dict(
        Products=products,
        IdleProducts=idle_products,
        IdleActiveByMonth=IdleActiveByMonth,
        MissingIdleConvCostProducts=missing_idle_conv_cost_products,
        Months=Months,
        R=R,
        RRound2=RRound2,
        A0=A0,
        A0Raw=A0Raw,
        IdleCapProd=IdleCapProd,
        ConvCost=ConvCost,
        AllowedConvPairs=AllowedConvPairs,
        NewBuyCost=NewBuyCost,
        month_key_map=month_map,
        month_keys=month_keys,
        month_label_by_key=month_label_by_key,
        month_label_by_index=month_label_by_index,
    )


# ----------------------------
# 2) Optimization Model
# ----------------------------
def solve_lca(
    data: dict,
    allocation_timing_mode: str = "requirement_month",
    lp_output_path: str | None = None,
):
    Products = data["Products"]
    IdleProducts = set(data["IdleProducts"])
    IdleActiveByMonth = data.get("IdleActiveByMonth", {})
    Months = data["Months"]
    R = data["R"]
    A0 = data["A0"]
    ConvCost = data["ConvCost"]
    AllowedConvPairs = data["AllowedConvPairs"]
    NewBuyCost = data["NewBuyCost"]
    IdleCapProd = data.get("IdleCapProd", {})

    solver = pywraplp.Solver.CreateSolver("SCIP")
    if solver is None:
        raise RuntimeError("SCIP solver not available.")

    valid_timing_modes = {"cost_optimal", "requirement_month"}
    if allocation_timing_mode not in valid_timing_modes:
        raise ValueError(
            f"allocation_timing_mode must be one of {sorted(valid_timing_modes)}; "
            f"got '{allocation_timing_mode}'"
        )

    # ---------------- Variables ----------------
    A, I, Sur, Short, IdleInv = {}, {}, {}, {}, {}
    x, y, Buy, z = {}, {}, {}, {}
    conv_to_dest_active = {}
    source_active = {}

    # Big-M for linking conversion quantity with destination selector.
    # Keep this safely large, based on total in-horizon supply + demand signals.
    total_requirement = sum(max(float(R.get((p, t), 0.0)), 0.0) for p in Products for t in Months)
    total_initial_inventory = sum(max(float(A0.get(p, 0.0)), 0.0) for p in Products)
    total_idle_capacity = sum(max(float(IdleCapProd.get((p, t), 0.0)), 0.0) for p in Products for t in Months)
    conversion_big_m = max(1.0, total_requirement + total_initial_inventory + total_idle_capacity)

    # Precompute conversion graph slices once for cleaner and faster constraint construction.
    dests_by_source = {
        q: [p for p in Products if p != q and (q, p) in AllowedConvPairs]
        for q in Products
    }
    srcs_by_dest = {
        p: [q for q in Products if q != p and (q, p) in AllowedConvPairs]
        for p in Products
    }

    for p in Products:
        for t in Months:
            A[(p, t)] = solver.IntVar(0, solver.infinity(), f"A_{p}_{t}")
            I[(p, t)] = solver.IntVar(0, solver.infinity(), f"I_{p}_{t}")
            Sur[(p, t)] = solver.IntVar(0, solver.infinity(), f"Sur_{p}_{t}")
            Short[(p, t)] = solver.IntVar(0, solver.infinity(), f"Short_{p}_{t}")
            IdleInv[(p, t)] = solver.IntVar(0, solver.infinity(), f"IdleInv_{p}_{t}")
            Buy[(p, t)] = solver.IntVar(0, solver.infinity(), f"Buy_{p}_{t}")
            is_idle_active = IdleActiveByMonth.get((p, t), p in IdleProducts)
            if is_idle_active:
                y[(p, t)] = solver.IntVar(0, solver.infinity(), f"y_{p}_{t}")
            else:
                y[(p, t)] = solver.IntVar(0, 0, f"y_{p}_{t}")
            conv_to_dest_active[(p, t)] = solver.BoolVar(f"conv_to_dest_active_{p}_{t}")

    for q in Products:
        for t in Months:
            source_active[(q, t)] = solver.BoolVar(f"src_active_{q}_{t}")
        for p in dests_by_source[q]:
            for t in Months:
                x[(p, q, t)] = solver.IntVar(0, solver.infinity(), f"x_{p}_{q}_{t}")
                z[(p, q, t)] = solver.BoolVar(f"z_{p}_{q}_{t}")

    # ---------------- Constraints ----------------
    # (0) Conversion into destination p is allowed only in months with requirement.
    # This prevents pre-conversion into months with zero demand.
    for p in Products:
        for t in Months:
            inflow_to_p_t = solver.Sum(x[(p, q, t)] for q in srcs_by_dest[p])
            required_qty = max(float(R.get((p, t), 0.0)), 0.0)
            if required_qty <= 1e-9:
                solver.Add(inflow_to_p_t == 0)
            else:
                solver.Add(inflow_to_p_t <= required_qty)

    # (1) Inventory definition per month:
    # Inventory is cumulative. Month t starts from month t-1 inventory,
    # then adjusts by conversion inflow completed in month t,
    # and source deduction for conversions that will complete in month t+1.
    # This means source inventory starts decreasing from one month before completion.
    # Buy follows 2-month lead time: order at t-2, receive in month t.
    for p in Products:
        for idx, t in enumerate(Months):
            inflow = solver.Sum(x[(p, q, t)] for q in srcs_by_dest[p])
            if idx < len(Months) - 1:
                t_next = Months[idx + 1]
                outflow = solver.Sum(x[(q, p, t_next)] for q in dests_by_source[p])
            else:
                outflow = 0
            if idx == 0:
                solver.Add(I[(p, t)] == A0[p] + inflow - outflow)
            elif idx == 1:
                t_prev = Months[idx - 1]
                solver.Add(I[(p, t)] == I[(p, t_prev)] + inflow - outflow)
            else:
                t_prev = Months[idx - 1]
                t_minus_2 = Months[idx - 2]
                solver.Add(I[(p, t)] == I[(p, t_prev)] + inflow - outflow + Buy[(p, t_minus_2)])

    # (2) Allocation definition:
    # Allocation = current month available inventory.
    for p in Products:
        for t in Months:
            solver.Add(A[(p, t)] == I[(p, t)])

    # (3) Surplus / Shortage from allocation vs requirement
    for p in Products:
        for t in Months:
            solver.Add(Sur[(p, t)] - Short[(p, t)] == A[(p, t)] - R[(p, t)])

    # (3.1) Buy must be ordered two months before the shortage month.
    # For month t shortage, order is Buy at month t-2.
    for p in Products:
        for idx, t in enumerate(Months):
            if idx <= len(Months) - 3:
                t_plus_2 = Months[idx + 2]
                solver.Add(Buy[(p, t)] == Short[(p, t_plus_2)])
            else:
                # Last two order months have no in-horizon shortage month to map.
                solver.Add(Buy[(p, t)] == 0)

    # Without pre-horizon buy orders, the first two months cannot have shortage.
    for p in Products:
        if len(Months) >= 1:
            solver.Add(Short[(p, Months[0])] == 0)
        if len(Months) >= 2:
            solver.Add(Short[(p, Months[1])] == 0)

    # (3.2) Shortage cannot exceed that month's requirement.
    # This avoids degenerate Sur/Short inflation when buy cost is zero.
    for p in Products:
        for t in Months:
            monthly_req = max(float(R.get((p, t), 0.0)), 0.0)
            solver.Add(Short[(p, t)] <= monthly_req)

    # Products that should be tracked in idle ledger.
    # Non-idle products are handled only by main inventory ledger I.
    idle_ledger_products = {
        p for p in Products
        if (p in IdleProducts) or any(IdleActiveByMonth.get((p, t), False) for t in Months)
    }

    # (4) Idle usage limit (current month)
    for p in Products:
        for t in Months:
            solver.Add(y[(p, t)] <= IdleInv[(p, t)])

    # (4.1) Idle inventory carry-over by reservation timing.
    # Idle inventory follows the same source-deduction timing as main inventory:
    # conversion finishing at month t+1 consumes source in month t.
    for p in Products:
        if p not in idle_ledger_products:
            for t in Months:
                solver.Add(IdleInv[(p, t)] == 0)
            continue

        for idx, t in enumerate(Months):
            conv_in_t = solver.Sum(x[(p, q, t)] for q in srcs_by_dest[p])
            if idx < len(Months) - 1:
                t_next = Months[idx + 1]
                conv_out_reserve_t = solver.Sum(x[(q, p, t_next)] for q in dests_by_source[p])
            else:
                conv_out_reserve_t = 0

            if idx == 0:
                solver.Add(
                    IdleInv[(p, t)]
                    == IdleCapProd[(p, t)] + conv_in_t - conv_out_reserve_t
                )
            else:
                t_prev = Months[idx - 1]
                solver.Add(
                    IdleInv[(p, t)]
                    == IdleInv[(p, t_prev)] + conv_in_t - conv_out_reserve_t
                )

    # (5) Conversion limit by surplus (2 months ago)
    for q in Products:
        for idx, t in enumerate(Months):
            if idx < 2:
                solver.Add(solver.Sum(x[(p, q, t)] for p in dests_by_source[q]) == 0)
            else:
                t_minus_2 = Months[idx - 2]
                solver.Add(
                    solver.Sum(x[(p, q, t)] for p in dests_by_source[q])
                    <= I[(q, t_minus_2)]
                )

    # (6) Idle usage limit (2 months ago)
    for p in Products:
        for idx, t in enumerate(Months):
            if idx < 2:
                solver.Add(y[(p, t)] == 0)
            else:
                t_minus_2 = Months[idx - 2]
                solver.Add(y[(p, t)] <= IdleCapProd[(p, t_minus_2)])

    # (7) Conversion pair activation.
    # Allow one source to feed multiple destinations in the same month,
    # but keep each destination limited to at most one source in that month.
    # Route selector z is active only when there is actual shipped quantity.
    for q in Products:
        destinations = dests_by_source[q]
        for t in Months:
            if not destinations:
                solver.Add(source_active[(q, t)] == 0)
                continue
            for p in destinations:
                solver.Add(x[(p, q, t)] <= conversion_big_m * z[(p, q, t)])
                solver.Add(x[(p, q, t)] >= z[(p, q, t)])
                solver.Add(z[(p, q, t)] <= source_active[(q, t)])
            solver.Add(source_active[(q, t)] <= solver.Sum(z[(p, q, t)] for p in destinations))

    for p in Products:
        sources = srcs_by_dest[p]
        for t in Months:
            if not sources:
                solver.Add(conv_to_dest_active[(p, t)] == 0)
                continue
            solver.Add(
                solver.Sum(z[(p, q, t)] for q in sources)
                <= conversion_big_m * conv_to_dest_active[(p, t)]
            )
            solver.Add(
                conv_to_dest_active[(p, t)]
                <= solver.Sum(z[(p, q, t)] for q in sources)
            )
            solver.Add(solver.Sum(z[(p, q, t)] for q in sources) <= 1)

    # Optional timing rule: when a destination receives conversion in a month,
    # do not allow surplus at that destination-month. This prevents pre-building
    # inventory via conversion before it is actually needed.
    if allocation_timing_mode == "requirement_month":
        for p in Products:
            for t in Months:
                solver.Add(Sur[(p, t)] <= conversion_big_m * (1 - conv_to_dest_active[(p, t)]))

    # (7.1) Source cannot convert in consecutive months (cooldown 1 month).
    # A source can feed multiple destinations in a month, but cannot be active
    # in both month t and t+1.
    for q in Products:
        for idx, t in enumerate(Months):
            if idx >= len(Months) - 1:
                continue
            t_next = Months[idx + 1]
            solver.Add(source_active[(q, t)] + source_active[(q, t_next)] <= 1)

    # ---------------- Objective ----------------
    # Pure cost-driven objective: choose the lowest total of buy + conversion costs.
    objective = solver.Sum(
        NewBuyCost[p] * Buy[(p, t)]
        for p in Products for t in Months
    )
    objective += solver.Sum(
        ConvCost[(q, p)] * x[(p, q, t)]
        for (q, p) in AllowedConvPairs for t in Months
    )
    solver.Minimize(objective)

    if lp_output_path:
        lp_text = solver.ExportModelAsLpFormat(False)
        with open(lp_output_path, "w", encoding="utf-8") as f:
            f.write(lp_text)

    status = solver.Solve()
    if status != pywraplp.Solver.OPTIMAL:
        status_name = _solver_status_name(status)
        if status == pywraplp.Solver.INFEASIBLE:
            diagnostics = _diagnose_infeasibility(data)
            detail = "\n - " + "\n - ".join(diagnostics)
            raise RuntimeError(
                f"No optimal solution. Status={status} ({status_name}). "
                f"Likely causes:{detail}"
            )
        raise RuntimeError(f"No optimal solution. Status={status} ({status_name})")

    return {
        "solver": solver,
        "A": A, "I": I, "Sur": Sur, "Short": Short,
        "x": x, "y": y, "IdleInv": IdleInv, "Buy": Buy,
        "objective_value": float(solver.Objective().Value()),
    }

# ----------------------------
# 3) Output (Excel per product)
# ----------------------------
def export_outputs(data, sol, output_path="LCA_Output.xlsx"):
    Products = data["Products"]
    Months = data["Months"]
    R = data["R"]
    A = sol["A"]
    I = sol["I"]
    Sur = sol["Sur"]
    Short = sol["Short"]
    RDisplay = data.get("RRound2", R)
    IdleInv = sol.get("IdleInv", {})
    x = sol["x"]

    (
        _,
        df_conversion,
        _,
        df_conversion_accounting_by_to_product,
        df_total_cost,
        df_shortage_detail,
    ) = build_usage_and_cost_reports(data, sol)

    month_label_by_index = data.get("month_label_by_index", {})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        (
            _,
            df_req_summary,
            _,
            _,
        ) = build_monthly_summary_tables(data, sol)

        # Write summary sheets first so they appear before detail and Product_* sheets.
        _write_visual_summary_like_sheet(
            writer=writer,
            data=data,
            sol=sol,
            df_conversion=df_conversion,
            df_conversion_accounting_by_to_product=df_conversion_accounting_by_to_product,
            df_total_cost=df_total_cost,
        )

        df_total_cost.to_excel(writer, sheet_name="Total_Cost_Summary", index=False)
        df_req_summary.to_excel(writer, sheet_name="LCA_Requirement_Summary", index=False)
        df_conversion.to_excel(writer, sheet_name="Conversion_Details", index=False)
        df_conversion_accounting_by_to_product.to_excel(writer, sheet_name="Conversion_Accounting_By_ToProduct", index=False)
        df_shortage_detail.to_excel(writer, sheet_name="Shortage_Details", index=False)

        for p in Products:
            cols = [month_label_by_index.get(t, f"M{t}") for t in Months]
            rows = ["LCA Inventory", "LCA Requirement", "Surplus", "Shortage", "LCA Conversion", "LCA Allocation"]
            table = {c: [0]*len(rows) for c in cols}
            sheet_name = _safe_sheet_name(f"Product_{p}")

            for idx, t in enumerate(Months):
                inv = _get_display_inventory_value(
                    product=p,
                    month_idx=t,
                    model_value=I[(p, t)].solution_value(),
                    data=data,
                )
                req = float(RDisplay.get((p, t), R[(p, t)]))
                surplus = Sur[(p, t)].solution_value()
                shortage = Short[(p, t)].solution_value()
                conv_in = sum(
                    x[(p, q, t)].solution_value()
                    for q in Products
                    if q != p and (p, q, t) in x
                )
                alloc = _get_display_inventory_value(
                    product=p,
                    month_idx=t,
                    model_value=A[(p, t)].solution_value(),
                    data=data,
                )

                col = cols[idx]
                table[col][0] = round(inv, 2)
                table[col][1] = round(req, 2)
                table[col][2] = round(surplus, 2)
                table[col][3] = round(shortage, 2)
                table[col][4] = round(conv_in, 2)
                table[col][5] = round(alloc, 2)

            df = pd.DataFrame(table, index=rows)
            df.index.name = f"Product {p}"
            df.to_excel(writer, sheet_name=sheet_name)

    print(f"Saved output to {output_path}")


# ----------------------------
# 4) Output (Excel formatting in summary sheet)
# ----------------------------
# excel formatting in summary sheet, with allocation split by idle vs non-idle when applicable, and showing requirement for reference.
def _write_visual_summary_like_sheet(
    writer,
    data,
    sol,
    df_conversion: pd.DataFrame,
    df_conversion_accounting_by_to_product: pd.DataFrame,
    df_total_cost: pd.DataFrame,
) -> None:
    Products = data["Products"]
    base_idle_product_set = set(data.get("IdleProducts", []))
    idle_active_by_month = data.get("IdleActiveByMonth", {})
    dynamic_idle_product_set = {
        p for p in Products
        if any(bool(idle_active_by_month.get((p, t), False)) for t in data["Months"])
    }
    # Show idle section as: base idle products + products that become idle-active by 3-zero-month rule.
    idle_product_set = base_idle_product_set.union(dynamic_idle_product_set)
    Months = data["Months"]
    R = data["R"]
    RRound2 = data.get("RRound2", R)
    A = sol["A"]
    month_label_by_index = data.get("month_label_by_index", {})
    month_columns = [month_label_by_index.get(t, f"M{t}") for t in Months]

    sheet_name = "Allocation_Summary"
    ws = writer.book.create_sheet(sheet_name)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Header row
    ws.cell(row=1, column=1, value="Type").font = header_font
    ws.cell(row=1, column=2, value="Product").font = header_font
    for idx, col_name in enumerate(month_columns, start=3):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = header_font
        cell.alignment = center

    requirement_products = [
        p for p in Products
        if any(abs(float(RRound2.get((p, t), 0.0))) > 1e-9 for t in Months)
    ]

    has_future_requirement = {
        (p, t): any(abs(float(R.get((p, tt), 0.0))) > 1e-9 for tt in Months if tt >= t)
        for p in Products
        for t in Months
    }

    def _is_idle_month(product: str, month_idx: int) -> bool:
        # A product can be shown as idle only when:
        # 1) it is idle-active by the zero-requirement rule (or base idle),
        # 2) it has inventory/allocation in that month,
        # 3) there is no remaining future requirement.
        has_inventory = abs(float(A[(product, month_idx)].solution_value())) > 1e-9
        if not has_inventory:
            return False
        if has_future_requirement.get((product, month_idx), False):
            return False
        return bool(idle_active_by_month.get((product, month_idx), product in idle_product_set))

    allocation_products = [
        p for p in Products
        if any(abs(A[(p, t)].solution_value()) > 1e-9 for t in Months)
    ]

    alloc_non_idle = [
        p for p in allocation_products
        if any((not _is_idle_month(p, t)) and abs(A[(p, t)].solution_value()) > 1e-9 for t in Months)
    ]
    alloc_idle = [
        p for p in allocation_products
        if any(_is_idle_month(p, t) and abs(A[(p, t)].solution_value()) > 1e-9 for t in Months)
    ]

    req_row_by_product: dict[str, int] = {}
    alloc_row_by_product_state: dict[tuple[str, str], int] = {}

    row = 2

    ws.cell(row=row, column=1, value="Requirement").font = header_font
    row += 1
    for p in requirement_products:
        ws.cell(row=row, column=1, value="Requirement")
        ws.cell(row=row, column=2, value=p)
        req_row_by_product[p] = row
        for idx, t in enumerate(Months, start=3):
            ws.cell(row=row, column=idx, value=round(float(RRound2.get((p, t), 0.0)), 4)).alignment = center
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Allocation (Non-Idle)").font = header_font
    row += 1
    for p in alloc_non_idle:
        ws.cell(row=row, column=1, value="Allocation")
        ws.cell(row=row, column=2, value=p)
        alloc_row_by_product_state[(p, "non_idle")] = row
        for idx, t in enumerate(Months, start=3):
            value = (
                round(
                    _get_display_inventory_value(
                        product=p,
                        month_idx=t,
                        model_value=A[(p, t)].solution_value(),
                        data=data,
                    ),
                    4,
                )
                if not _is_idle_month(p, t)
                else None
            )
            ws.cell(row=row, column=idx, value=value).alignment = center
        row += 1

    if alloc_idle:
        row += 1
        ws.cell(row=row, column=1, value="Allocation (Idle)").font = header_font
        row += 1
        for p in alloc_idle:
            ws.cell(row=row, column=1, value="Allocation")
            ws.cell(row=row, column=2, value=p)
            alloc_row_by_product_state[(p, "idle")] = row
            for idx, t in enumerate(Months, start=3):
                value = (
                    round(
                        _get_display_inventory_value(
                            product=p,
                            month_idx=t,
                            model_value=A[(p, t)].solution_value(),
                            data=data,
                        ),
                        4,
                    )
                    if _is_idle_month(p, t)
                    else None
                )
                ws.cell(row=row, column=idx, value=value).alignment = center
            row += 1

    row += 2
    ws.cell(row=row, column=1, value="Conversion_Accounting").font = header_font
    row += 1
    for _, acc_row in df_conversion_accounting_by_to_product.iterrows():
        product_name = str(acc_row.get("Product", "")).strip()
        if not product_name:
            continue
        ws.cell(row=row, column=1, value="Conversion_Accounting")
        ws.cell(row=row, column=2, value=product_name)
        for idx, month_label in enumerate(month_columns, start=3):
            value = acc_row.get(month_label, 0.0)
            try:
                numeric_value = float(value)
            except Exception:
                numeric_value = 0.0
            ws.cell(row=row, column=idx, value=round(numeric_value, 4)).alignment = center
        row += 1

    cost_header_row = row + 2
    ws.cell(row=cost_header_row, column=1, value="Cost").font = header_font
    ws.cell(row=cost_header_row, column=2, value="Flow").font = header_font
    ws.cell(row=cost_header_row, column=3, value="TotalQty").font = header_font
    ws.cell(row=cost_header_row, column=4, value="UnitCost").font = header_font
    ws.cell(row=cost_header_row, column=5, value="TotalCost").font = header_font

    pair_to_color: dict[tuple[str, str], str] = {}
    if not df_conversion.empty:
        flow_cost = (
            df_conversion.groupby(["FromProduct", "ToProduct"], as_index=False)[["ConversionQty", "ConvCost"]]
            .sum()
            .rename(columns={"ConversionQty": "TotalQty", "ConvCost": "TotalCost"})
            .sort_values(["TotalCost", "FromProduct", "ToProduct"], ascending=[False, True, True])
            .reset_index(drop=True)
        )

        for idx, flow in flow_cost.iterrows():
            pair = (str(flow["FromProduct"]), str(flow["ToProduct"]))
            pair_to_color[pair] = _FLOW_COLOR_PALETTE[idx % len(_FLOW_COLOR_PALETTE)]

        for idx, flow in flow_cost.iterrows():
            out_row = cost_header_row + 1 + idx
            pair = (str(flow["FromProduct"]), str(flow["ToProduct"]))
            qty = float(flow["TotalQty"])
            total_cost = float(flow["TotalCost"])
            unit_cost = (total_cost / qty) if abs(qty) > 1e-9 else 0.0
            ws.cell(row=out_row, column=1, value="Cost")
            ws.cell(row=out_row, column=2, value=f"{pair[0]} -> {pair[1]}")
            ws.cell(row=out_row, column=3, value=round(qty, 4)).alignment = center
            ws.cell(row=out_row, column=4, value=round(unit_cost, 4)).alignment = center
            ws.cell(row=out_row, column=5, value=round(total_cost, 4)).alignment = center
            fill = PatternFill(fill_type="solid", fgColor=pair_to_color[pair])
            for col in range(1, 6):
                ws.cell(row=out_row, column=col).fill = fill

    # Color highlight by conversion flow month, similar to the visual sample.
    month_to_col = {label: idx for idx, label in enumerate(month_columns, start=3)}
    month_order_by_label = {
        str(month_label_by_index.get(t, f"M{t}")): t
        for t in Months
    }
    if not df_conversion.empty:
        for _, conv in df_conversion.iterrows():
            pair = (str(conv["FromProduct"]), str(conv["ToProduct"]))
            color_hex = pair_to_color.get(pair)
            if not color_hex:
                continue
            fill = PatternFill(fill_type="solid", fgColor=color_hex)
            month_label = str(conv["Month"])
            col_to = month_to_col.get(month_label)
            if col_to is None:
                continue

            t_to = month_order_by_label.get(month_label)
            if t_to is None:
                continue

            from_p, to_p = pair
            to_state = "idle" if _is_idle_month(to_p, t_to) else "non_idle"
            req_row = req_row_by_product.get(to_p)
            alloc_to_row = alloc_row_by_product_state.get((to_p, to_state))

            if req_row is not None:
                ws.cell(row=req_row, column=col_to).fill = fill
            if alloc_to_row is not None:
                ws.cell(row=alloc_to_row, column=col_to).fill = fill
            if from_p:
                t_from = t_to - 2
                if t_from < 1:
                    continue
                from_month_label = str(month_label_by_index.get(t_from, f"M{t_from}"))
                col_from = month_to_col.get(from_month_label)
                if col_from is None:
                    continue
                from_state = "idle" if _is_idle_month(from_p, t_from) else "non_idle"
                alloc_from_row = alloc_row_by_product_state.get((from_p, from_state))
                if alloc_from_row is not None:
                    ws.cell(row=alloc_from_row, column=col_from).fill = fill

    # Total cost line under cost table.
    cost_total_row = cost_header_row + 2
    if not df_conversion.empty:
        cost_total_row = cost_header_row + 1 + len(
            df_conversion.groupby(["FromProduct", "ToProduct"]).size()
        ) + 1
    total_cost_value = float(df_total_cost.loc[df_total_cost["CostType"] == "TotalCost", "Amount"].sum())
    ws.cell(row=cost_total_row, column=4, value="Total").font = header_font
    ws.cell(row=cost_total_row, column=5, value=round(total_cost_value, 4)).font = header_font

    ws.freeze_panes = "C2"


def build_usage_and_cost_reports(data, sol):
    Products = data["Products"]
    IdleProducts = set(data["IdleProducts"])
    IdleActiveByMonth = data.get("IdleActiveByMonth", {})
    Months = data["Months"]
    month_label_by_index = data.get("month_label_by_index", {})
    NewBuyCost = data["NewBuyCost"]
    ConvCost = data["ConvCost"]
    AllowedConvPairs = data["AllowedConvPairs"]

    def _period_label(month_idx: int) -> str:
        return str(month_label_by_index.get(month_idx, f"M{month_idx}"))

    x = sol["x"]
    IdleInv = sol.get("IdleInv", {})
    Short = sol["Short"]

    idle_rows = []
    for p in Products:
        if p not in IdleProducts and not any(IdleActiveByMonth.get((p, t), False) for t in Months):
            continue
        for t in Months:
            if p not in IdleProducts and not IdleActiveByMonth.get((p, t), False):
                continue
            if (p, t) in IdleInv:
                idle_inven_qty = IdleInv[(p, t)].solution_value()
            else:
                idle_inven_qty = float(data["IdleCapProd"].get((p, t), 0.0))
            conv_in_qty = sum(
                x[(p, q, t)].solution_value()
                for q in Products
                if q != p and (q, p) in AllowedConvPairs
            )
            conv_out_qty = sum(
                x[(q, p, t)].solution_value()
                for q in Products
                if q != p and (p, q) in AllowedConvPairs
            )
            idle_usage_qty = conv_in_qty - conv_out_qty
            # idle_qty_used_in_model = y[(p, t)].solution_value()

            if (
                abs(idle_inven_qty) > 1e-9
                or abs(conv_in_qty) > 1e-9
                or abs(conv_out_qty) > 1e-9
                # or abs(idle_qty_used_in_model) > 1e-9
            ):
                idle_rows.append(
                    {
                        "Product": p,
                        "Month": _period_label(t),
                        "_MonthOrder": t,
                        "IdleInvenQty": round(idle_inven_qty, 4),
                        "IdleUsageQty": round(idle_usage_qty, 4),
                        "ConvInQty": round(conv_in_qty, 4),
                        "ConvOutQty": round(conv_out_qty, 4),
                        # "IdleQtyUsedInModel": round(idle_qty_used_in_model, 4), 
                    }
                )
    df_idle_usage = pd.DataFrame(idle_rows)
    if not df_idle_usage.empty:
        df_idle_usage = df_idle_usage.sort_values(["_MonthOrder", "Product"]).reset_index(drop=True)
        df_idle_usage = df_idle_usage.drop(columns=["_MonthOrder"])

    conv_rows = []
    for p in Products:  # destination
        for q in Products:  # source
            if p == q or (q, p) not in AllowedConvPairs:
                continue
            for t in Months:
                qty = x[(p, q, t)].solution_value()
                if abs(qty) > 1e-9:
                    unit_cost = float(ConvCost.get((q, p), 0.0))
                    conv_rows.append(
                        {
                            "FromProduct": q,
                            "ToProduct": p,
                            "Month": _period_label(t),
                            "_MonthOrder": t,
                            "ConversionQty": round(qty, 4),
                            "ConvUnitCost": unit_cost,
                            "ConvCost": round(unit_cost * qty, 4),
                        }
                    )
    df_conversion = pd.DataFrame(conv_rows)
    if not df_conversion.empty:
        df_conversion = df_conversion.sort_values(["_MonthOrder", "FromProduct", "ToProduct"]).reset_index(drop=True)
        df_conversion = df_conversion.drop(columns=["_MonthOrder"])

    conv_accounting_rows = []
    for row in conv_rows:
        source_month_order = int(row["_MonthOrder"])
        accounting_month_order = source_month_order - 2
        conv_accounting_rows.append(
            {
                "ToProduct": row["ToProduct"],
                "AccountingMonth": _period_label(accounting_month_order) if accounting_month_order >= 1 else f"PreHorizon(M{accounting_month_order})",
                "SourceConversionMonth": row["Month"],
                "_AccountingMonthOrder": accounting_month_order,
                "_SourceConversionMonthOrder": source_month_order,
                "AccountingInHorizon": accounting_month_order >= 1,
                "ConversionQty": row["ConversionQty"],
                "ConvUnitCost": row["ConvUnitCost"],
                "ConvCost": row["ConvCost"],
            }
        )

    df_conversion_accounting_detail = pd.DataFrame(conv_accounting_rows)
    if not df_conversion_accounting_detail.empty:
        df_conversion_accounting_detail = df_conversion_accounting_detail.sort_values(
            ["_AccountingMonthOrder", "ToProduct", "_SourceConversionMonthOrder"]
        ).reset_index(drop=True)
        df_conversion_accounting_detail = df_conversion_accounting_detail.drop(
            columns=["_AccountingMonthOrder", "_SourceConversionMonthOrder"]
        )

        accounting_month_columns = [
            _period_label(t)
            for t in Months
        ]
        df_conversion_accounting_by_to_product = (
            df_conversion_accounting_detail
            .pivot_table(index="ToProduct", columns="AccountingMonth", values="ConvCost", aggfunc="sum", fill_value=0.0)
            .reindex(columns=accounting_month_columns, fill_value=0.0)
            .reset_index()
        )
        for col in accounting_month_columns:
            if col not in df_conversion_accounting_by_to_product.columns:
                df_conversion_accounting_by_to_product[col] = 0.0
            df_conversion_accounting_by_to_product[col] = (
                df_conversion_accounting_by_to_product[col].astype(float).round(4)
            )
        df_conversion_accounting_by_to_product = df_conversion_accounting_by_to_product[
            ["ToProduct"] + accounting_month_columns
        ]
    else:
        accounting_month_columns = [_period_label(t) for t in Months]
        df_conversion_accounting_by_to_product = pd.DataFrame(columns=["ToProduct"] + accounting_month_columns)
    df_conversion_accounting_by_to_product.rename(columns={"ToProduct": "Product"}, inplace=True)

    total_short_cost = sum(
        float(NewBuyCost.get(p, 0.0)) * Short[(p, t)].solution_value()
        for p in Products for t in Months
    )
    total_conv_cost = sum(
        float(ConvCost.get((q, p), 0.0)) * x[(p, q, t)].solution_value()
        for (q, p) in AllowedConvPairs for t in Months
    )
    total_cost = total_short_cost + total_conv_cost

    df_total_cost = pd.DataFrame(
        [
            {"CostType": "PurchaseCost", "Amount": round(total_short_cost, 4)},
            {"CostType": "ConversionCost", "Amount": round(total_conv_cost, 4)},
            {"CostType": "TotalCost", "Amount": round(total_cost, 4)},
        ]
    )

    shortage_rows = []
    for p in Products:
        unit_cost = float(NewBuyCost.get(p, 0.0))
        for t in Months:
            qty = Short[(p, t)].solution_value()
            if abs(qty) > 1e-9:
                shortage_rows.append(
                    {
                        "Product": p,
                        "Month": _period_label(t),
                        "_MonthOrder": t,
                        "ShortageQty": round(qty, 4),
                        "PurchaseUnitCost": unit_cost,
                        "ShortageCost": round(unit_cost * qty, 4),
                    }
                )
    df_shortage_detail = pd.DataFrame(shortage_rows)
    if not df_shortage_detail.empty:
        df_shortage_detail = df_shortage_detail.sort_values(["_MonthOrder", "Product"]).reset_index(drop=True)
        df_shortage_detail = df_shortage_detail.drop(columns=["_MonthOrder"])

    return (
        df_idle_usage,
        df_conversion,
        df_conversion_accounting_detail,
        df_conversion_accounting_by_to_product,
        df_total_cost,
        df_shortage_detail,
    )


def build_monthly_summary_tables(data, sol):
    Products = data["Products"]
    IdleProducts = _ordered_unique(data.get("IdleProducts", []))
    Months = data["Months"]
    month_label_by_index = data.get("month_label_by_index", {})
    R = data["R"]
    RDisplay = data.get("RRound2", R)

    # Show allocation/requirement summaries only for products with actual requirement.
    summary_products = [
        p for p in Products
        if any(abs(float(R.get((p, t), 0.0))) > 1e-9 for t in Months)
    ]

    A = sol["A"]
    IdleInv = sol.get("IdleInv", {})

    month_columns = [month_label_by_index.get(t, f"M{t}") for t in Months]

    alloc_rows = []
    req_rows = []
    idle_rows = []
    idle_usage_rows = []
    for p in summary_products:
        alloc_row = {"Product": p}
        req_row = {"Product": p}
        idle_row = {"Product": p}
        for t in Months:
            col = month_label_by_index.get(t, f"M{t}")
            alloc_row[col] = round(
                _get_display_inventory_value(
                    product=p,
                    month_idx=t,
                    model_value=A[(p, t)].solution_value(),
                    data=data,
                ),
                4,
            )
            req_row[col] = round(float(RDisplay.get((p, t), R.get((p, t), 0.0))), 4)
            idle_value = IdleInv[(p, t)].solution_value() if (p, t) in IdleInv else float(data["IdleCapProd"].get((p, t), 0.0))
            idle_row[col] = round(idle_value, 4)
        alloc_rows.append(alloc_row)
        req_rows.append(req_row)
        idle_rows.append(idle_row)

    # Build Idle Usage summary explicitly from IdleInvenQty in detailed usage table.
    df_idle_usage_detail, *_ = build_usage_and_cost_reports(data, sol)
    if not df_idle_usage_detail.empty:
        usage_pivot = (
            df_idle_usage_detail
            .pivot_table(index="Product", columns="Month", values="IdleInvenQty", aggfunc="sum", fill_value=0.0)
            .reindex(index=IdleProducts, columns=month_columns, fill_value=0.0)
        )
        usage_pivot = usage_pivot.reset_index()
        for col in month_columns:
            if col not in usage_pivot.columns:
                usage_pivot[col] = 0.0
        df_idle_usage_summary = usage_pivot[["Product"] + month_columns].copy()
        for col in month_columns:
            df_idle_usage_summary[col] = df_idle_usage_summary[col].astype(float).round(4)
    else:
        for p in IdleProducts:
            idle_usage_row = {"Product": p}
            for col in month_columns:
                idle_usage_row[col] = 0.0
            idle_usage_rows.append(idle_usage_row)
        df_idle_usage_summary = pd.DataFrame(idle_usage_rows, columns=["Product"] + month_columns)

    df_alloc_summary = pd.DataFrame(alloc_rows, columns=["Product"] + month_columns)
    df_req_summary = pd.DataFrame(req_rows, columns=["Product"] + month_columns)
    df_idle_inventory_summary = pd.DataFrame(idle_rows, columns=["Product"] + month_columns)
    return df_alloc_summary, df_req_summary, df_idle_inventory_summary, df_idle_usage_summary


# ----------------------------
# Main
# ----------------------------
if __name__ == "__main__":
    script_start = time.perf_counter()
    
    ## Example usage with demo input file
    path = r"C:\Users\768940\OneDrive - Seagate Technology\Documents\My Task\DS I\knime_python\HSA\POR_HSA_AI_in_OS\workshop\input\Demo_input_POR4.xlsx"
    now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = fr"C:\Users\768940\OneDrive - Seagate Technology\Documents\My Task\DS I\knime_python\HSA\POR_HSA_AI_in_OS\workshop\output\LCA_Output_POR4_{now_str}.xlsx"
    lp_output_path = fr"C:\Users\768940\OneDrive - Seagate Technology\Documents\My Task\DS I\knime_python\HSA\POR_HSA_AI_in_OS\workshop\output\LCA_Model_POR4_{now_str}.lp"

    data = build_data_from_demo_input(path=path)

    if data.get("MissingIdleConvCostProducts"):
        print("Warning: Some idle products have no conversion cost pairs in df_conv_cost:")
        print(data["MissingIdleConvCostProducts"])

    solve_start = time.perf_counter()
    sol = solve_lca(data, lp_output_path=lp_output_path)
    solve_elapsed = time.perf_counter() - solve_start
    print(f"\n=== Optimization Solve Time ===\n{solve_elapsed:.4f} seconds")
    print(f"\n=== LP Model Saved ===\n{lp_output_path}")

    export_outputs(data, sol, output_path=output_path)

    (
        df_alloc_summary,
        df_req_summary,
        df_idle_inventory_summary,
        df_idle_usage_summary,
    ) = build_monthly_summary_tables(data, sol)
    print("\n=== LCA Allocation Summary (All Products) ===")
    print(df_alloc_summary.to_string(index=False))
    print("\n=== LCA Requirement Summary (All Products) ===")
    print(df_req_summary.to_string(index=False))
    print("\n=== Idle Usage Summary (IdleInvenQty, All Products) ===")
    print(df_idle_usage_summary.to_string(index=False))

    target_product = "Marlin / MarlinBP / Summit"
    if target_product not in data["Products"]:
        print(f"Product '{target_product}' not found in input. Available products: {data['Products']}")
    else:
        Products = data["Products"]
        Months = data["Months"]
        R = data["R"]
        RDisplay = data.get("RRound2", R)
        A = sol["A"]
        I = sol["I"]
        Sur = sol["Sur"]
        Short = sol["Short"]
        x = sol["x"]

        month_label_by_index = data.get("month_label_by_index", {})
        cols = [month_label_by_index.get(t, f"M{t}") for t in Months]
        rows = ["LCA Inventory", "LCA Requirement", "Surplus", "Shortage", "LCA Conversion", "LCA Allocation"]
        table = {c: [0] * len(rows) for c in cols}

        for idx, t in enumerate(Months):
            inv = _get_display_inventory_value(
                product=target_product,
                month_idx=t,
                model_value=I[(target_product, t)].solution_value(),
                data=data,
            )
            req = float(RDisplay.get((target_product, t), R[(target_product, t)]))
            surplus = Sur[(target_product, t)].solution_value()
            shortage = Short[(target_product, t)].solution_value()
            conv_in = sum(
                x[(target_product, q, t)].solution_value()
                for q in Products
                if q != target_product and (target_product, q, t) in x
            )
            alloc = _get_display_inventory_value(
                product=target_product,
                month_idx=t,
                model_value=A[(target_product, t)].solution_value(),
                data=data,
            )

            col = cols[idx]
            table[col][0] = round(inv, 2)
            table[col][1] = round(req, 2)
            table[col][2] = round(surplus, 2)
            table[col][3] = round(shortage, 2)
            table[col][4] = round(conv_in, 2)
            table[col][5] = round(alloc, 2)

        df = pd.DataFrame(table, index=rows)
        df.index.name = f"Product {target_product}"
        print(df)

    (
        df_idle_usage,
        df_conversion,
        df_conversion_accounting_detail,
        df_conversion_accounting_by_to_product,
        df_total_cost,
        df_shortage_detail,
    ) = build_usage_and_cost_reports(data, sol)

    print("\n=== Idle Usage ===")
    if df_idle_usage.empty:
        print("No idle usage.")
    else:
        print(df_idle_usage.to_string(index=False))

    print("\n=== Conversion Details ===")
    if df_conversion.empty:
        print("No conversion.")
    else:
        print(df_conversion.to_string(index=False))

    print("\n=== Conversion Accounting by ToProduct (Month-2) ===")
    if df_conversion_accounting_by_to_product.empty:
        print("No conversion accounting rows.")
    else:
        print(df_conversion_accounting_by_to_product.to_string(index=False))

    print("\n=== Shortage Details (Product/Month) ===")
    if df_shortage_detail.empty:
        print("No shortage.")
    else:
        print(df_shortage_detail.to_string(index=False))
    
    print("\n=== Total Cost Summary ===")
    print(df_total_cost.to_string(index=False))

    script_elapsed = time.perf_counter() - script_start
    print(f"\n=== Total Script Runtime ===\n{script_elapsed:.4f} seconds")